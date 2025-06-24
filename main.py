from openpyxl import load_workbook
from datetime import datetime
now = datetime.now()
months_uk = {
    1: 'січень',
    2: 'лютий',
    3: 'березень',
    4: 'квітень',
    5: 'травень',
    6: 'червень',
    7: 'липень',
    8: 'серпень',
    9: 'вересень',
    10: 'жовтень',
    11: 'листопад',
    12: 'грудень',
}

auto = "str"
trailer = "str"

driver = "str" #(П. І. Б., номер посвідчення водія)

ttn_number = "str" # номер ТТН
name_of_goods = "str" # товар

customer = "str" # Едрпоу адреса назва компании 
shipper = "str" # повне найменування, місцезнаходження / П. І.Б., місце проживання
consignee = "str" # повне найменування, місцезнаходження / П. І.Б., місце проживання

point_of_loading = "str" # місце завантаження
point_of_unloading = "str" # місце вивантаження



day = now.day
data = {
    'Y4': ttn_number, # номер ТТН
    'AG4': day, # день
    'AJ4': months_uk[now.month],
    'I6':  auto, # (марка, модель, тип, реєстраційний номер)
    'AE6': trailer, # (марка, модель, тип, реєстраційний номер
    'F38': name_of_goods, # товар
    'L10': customer, # замовник
    'L11': shipper, # Вантажовідправник
    'L15': consignee, # Вантажовідправник
    'L17': point_of_loading, # місце завантаження
    'AO17': point_of_unloading, # місце вивантаження
    'AS8': driver, # Водій
}

wb = load_workbook('inbulk_ttn.xlsx')
ws = wb.active

for cell, value in data.items():
    try:
        ws[cell] = value
    except AttributeError:
        print(f"Не удалось записать в {cell}: объединённая ячейка. Запишите в верхнюю левую ячейку диапазона.")

wb.save('filled_ttn2.xlsx')